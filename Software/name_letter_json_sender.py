#!/usr/bin/env python3

import json
import time
from openpyxl import load_workbook


class NameLetterJsonSender:
    def __init__(
        self,
        middleware,
        excel_path="asl_static_right_hand_dataset.xlsx"
    ):
        self.middleware = middleware
        self.excel_path = excel_path
        self.letter_map = self.load_letter_map()

    def load_letter_map(self):
        workbook = load_workbook(self.excel_path, data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        letter_map = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            label = row[0]

            if label is None:
                continue

            label = str(label).strip().upper()

            values = {
                headers[i]: row[i]
                for i in range(1, len(headers))
            }

            letter_map[label] = values

        print("[NameLetter] Excel letter map loaded.")
        return letter_map

    def process_json(self, json_text):
        try:
            data = json.loads(json_text)
        except json.JSONDecodeError:
            print("[NameLetter] Invalid JSON received.")
            return None

        gloss = data.get("gloss", "")
        names = data.get("names", [])

        output = {
            "gloss": gloss,
            "names": []
        }

        for name in names:
            name_data = {
                "name": name,
                "letters": []
            }

            for letter in name:
                letter_upper = letter.upper()

                if letter_upper in self.letter_map:
                    name_data["letters"].append({
                        "letter": letter_upper,
                        "values": self.letter_map[letter_upper]
                    })

            output["names"].append(name_data)

        output_json = json.dumps(output)

        print(f"[NameLetter] Output JSON: {output_json}")

        self.middleware.set_name_letters_output(output_json)

        return output_json

    def run_loop(self, poll_interval=0.2):
        print("[NameLetter] Running. Waiting for language JSON...")

        while True:
            json_text = self.middleware.get_language_output(clear=True)

            if json_text:
                self.process_json(json_text)

            time.sleep(poll_interval)


def main():
    from middleware import Middleware

    middleware = Middleware()

    sender = NameLetterJsonSender(
        middleware=middleware,
        excel_path="asl_static_right_hand_dataset.xlsx"
    )

    try:
        sender.run_loop()

    except KeyboardInterrupt:
        print("\n[NameLetter] Stopped by user.")


if __name__ == "__main__":
    main()